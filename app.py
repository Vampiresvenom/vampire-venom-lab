"""
╔══════════════════════════════════════════════════════════════╗
║  VAMPIRE VENOM LAB  ·  Enterprise BI Intelligence Platform   ║
║  Version 2.0  ·  Full Production Build                       ║
╚══════════════════════════════════════════════════════════════╝

Features:
  ✓ Login / Register / Session auth
  ✓ Free trial + Starter + Pro + Enterprise tiers
  ✓ Per-user report counting & enforcement
  ✓ Dashboard: report history, usage stats, billing, team
  ✓ Team / colleague access sharing
  ✓ Admin panel: manage users, override API key
  ✓ API key from env OR admin override (both supported)
  ✓ Sidebar always visible (CSS locked)
  ✓ Dark crimson-gold luxury aesthetic
  ✓ Full BI pipeline: analyze → clean → DAX → PQ → charts → export
"""

import streamlit as st
import pandas as pd
import numpy as np
import json, re, io, textwrap, hashlib, os, time, uuid
from datetime import datetime, timedelta
from copy import deepcopy

# ── Must be absolute first Streamlit call ────────────────────────────────────
st.set_page_config(
    page_title="Vampire Venom Lab",
    page_icon="🧪",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ════════════════════════════════════════════════════════════════
# ░░  GLOBAL CSS  ░░
# ════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cinzel+Decorative:wght@400;700&family=Cinzel:wght@400;600;700&family=Outfit:wght@300;400;500;600;700&display=swap');

/* ── Variables ── */
:root {
    --crimson:  #C41E3A;
    --crimson2: #8B0000;
    --gold:     #D4A017;
    --gold2:    #F5D76E;
    --bg:       #07090F;
    --bg2:      #0C0F1A;
    --card:     #0F1420;
    --card2:    #141B2D;
    --card3:    #1A2240;
    --border:   rgba(196,30,58,0.18);
    --border2:  rgba(212,160,23,0.18);
    --text:     #E8E2D9;
    --muted:    #5A6070;
    --success:  #10B981;
    --warn:     #F59E0B;
    --danger:   #EF4444;
}

/* ── Base ── */
* { font-family: 'Outfit', sans-serif; box-sizing: border-box; }
.cinzel { font-family: 'Cinzel', serif !important; }
html, body, .stApp { background: var(--bg) !important; color: var(--text) !important; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 1.5rem 2rem 5rem 2rem !important; max-width: 1280px !important; }

/* ── Sidebar: always visible, never collapsible ── */
[data-testid="stSidebar"] {
    background: linear-gradient(170deg, #0A0D18 0%, #07090F 100%) !important;
    border-right: 1px solid var(--border) !important;
    min-width: 270px !important;
    max-width: 270px !important;
}
[data-testid="stSidebar"] * { color: var(--text) !important; }
[data-testid="collapsedControl"],
button[kind="header"] { display: none !important; }

/* ── Primary buttons ── */
.stButton > button {
    background: linear-gradient(135deg, #C41E3A 0%, #7A0010 100%) !important;
    color: var(--gold2) !important;
    border: 1px solid rgba(212,160,23,0.35) !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-size: .93rem !important;
    padding: .62rem 1.4rem !important;
    letter-spacing: .3px !important;
    transition: all .25s ease !important;
    box-shadow: 0 0 18px rgba(196,30,58,.28) !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #E02040 0%, #9A0020 100%) !important;
    box-shadow: 0 0 32px rgba(196,30,58,.52) !important;
    transform: translateY(-1px) !important;
}

/* ── Download buttons ── */
.stDownloadButton > button {
    background: var(--card2) !important;
    color: var(--gold2) !important;
    border: 1px solid var(--border2) !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-size: .88rem !important;
    transition: all .25s !important;
}
.stDownloadButton > button:hover {
    background: rgba(212,160,23,.08) !important;
    border-color: var(--gold) !important;
    box-shadow: 0 0 16px rgba(212,160,23,.2) !important;
}

/* ── Inputs ── */
.stTextInput > div > div > input,
.stTextArea > div > textarea,
.stSelectbox > div > div > div {
    background: var(--card) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    color: var(--text) !important;
    font-family: 'Outfit', sans-serif !important;
    font-size: .93rem !important;
}
.stTextInput > div > div > input:focus { border-color: var(--crimson) !important; box-shadow: 0 0 14px rgba(196,30,58,.22) !important; }

/* ── Metrics ── */
[data-testid="stMetric"] { background: var(--card) !important; border: 1px solid var(--border) !important; border-radius: 12px !important; padding: 1rem 1.1rem !important; }
[data-testid="stMetricValue"] { color: var(--gold2) !important; font-weight: 700 !important; font-size: 1.5rem !important; }
[data-testid="stMetricLabel"] { color: var(--muted) !important; font-size: .78rem !important; text-transform: uppercase; letter-spacing: .5px; }
[data-testid="stMetricDelta"] { font-size: .8rem !important; }

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] { background: var(--card) !important; border-radius: 10px !important; padding: 4px !important; border: 1px solid var(--border) !important; gap: 3px !important; }
.stTabs [data-baseweb="tab"] { border-radius: 7px !important; color: var(--muted) !important; font-weight: 500 !important; font-size: .88rem !important; }
.stTabs [aria-selected="true"] { background: rgba(196,30,58,.2) !important; color: var(--gold2) !important; font-weight: 700 !important; }

/* ── Expander ── */
[data-testid="stExpander"] { background: var(--card) !important; border: 1px solid var(--border) !important; border-radius: 10px !important; margin-bottom: .5rem !important; }
[data-testid="stExpander"]:hover { border-color: var(--border2) !important; }

/* ── File uploader ── */
[data-testid="stFileUploader"] { background: var(--card) !important; border: 2px dashed rgba(196,30,58,.3) !important; border-radius: 14px !important; transition: all .3s !important; }
[data-testid="stFileUploader"]:hover { border-color: var(--gold) !important; }

/* ── Dataframe ── */
[data-testid="stDataFrame"] { border-radius: 10px !important; overflow: hidden !important; }

/* ── Progress bar ── */
.stProgress > div > div { background: linear-gradient(90deg, var(--crimson), var(--gold)) !important; border-radius: 99px !important; }

/* ── Checkbox ── */
.stCheckbox { accent-color: var(--crimson) !important; }

/* ── Divider ── */
hr { border-color: var(--border) !important; margin: 1rem 0 !important; }

/* ── Alerts ── */
.stAlert { border-radius: 10px !important; border-left-width: 3px !important; }

/* ── Spinner ── */
.stSpinner > div { border-top-color: var(--crimson) !important; }

/* ── Custom card components ── */
.vvl-card          { background: var(--card);  border: 1px solid var(--border);  border-radius: 14px; padding: 1.2rem 1.4rem; margin: .4rem 0; }
.vvl-card2         { background: var(--card2); border: 1px solid var(--border);  border-radius: 14px; padding: 1.2rem 1.4rem; margin: .4rem 0; }
.vvl-card-crimson  { border-left: 3px solid var(--crimson); }
.vvl-card-gold     { border-left: 3px solid var(--gold); }
.vvl-card-success  { border-left: 3px solid #10B981; }
.vvl-card-warn     { border-left: 3px solid #F59E0B; border-color: rgba(245,158,11,.2); }

/* ── Labels ── */
.vvl-label { font-family: 'Cinzel', serif; font-size: .68rem; color: var(--gold); letter-spacing: 2.5px; text-transform: uppercase; opacity: .85; margin-bottom: .55rem; display: block; }
.vvl-section { font-family: 'Cinzel', serif; font-size: 1.1rem; font-weight: 600; color: var(--text); margin: 1.5rem 0 .8rem 0; }

/* ── DAX code block ── */
.dax-block { background: #060810; border: 1px solid var(--border); border-radius: 10px; padding: 1rem 1.2rem; font-family: 'Courier New', monospace; font-size: .82rem; line-height: 1.75; color: var(--gold2); overflow-x: auto; white-space: pre-wrap; }

/* ── Plan badges ── */
.plan-badge { display: inline-block; padding: 2px 12px; border-radius: 99px; font-size: .7rem; font-weight: 700; letter-spacing: 1px; text-transform: uppercase; }
.badge-free       { background: rgba(90,96,112,.15);  color: #94A3B8; border: 1px solid rgba(90,96,112,.3); }
.badge-starter    { background: rgba(16,185,129,.1);  color: #34D399; border: 1px solid rgba(16,185,129,.25); }
.badge-pro        { background: rgba(212,160,23,.12); color: #F5D76E; border: 1px solid rgba(212,160,23,.3); }
.badge-enterprise { background: rgba(196,30,58,.12);  color: #F87171; border: 1px solid rgba(196,30,58,.3); }
.badge-admin      { background: rgba(139,0,0,.2);     color: #FCA5A5; border: 1px solid rgba(139,0,0,.4); }

/* ── Usage bar ── */
.usage-bar-bg  { background: var(--card2); border-radius: 99px; height: 7px; overflow: hidden; margin: 6px 0 2px 0; }
.usage-bar-fill { height: 100%; border-radius: 99px; transition: width .6s ease; }

/* ── Team member row ── */
.team-row { display: flex; align-items: center; justify-content: space-between; padding: .7rem 1rem; border-radius: 8px; border: 1px solid var(--border); background: var(--card2); margin-bottom: .4rem; }

/* ── Pricing card highlight ── */
.plan-popular { box-shadow: 0 0 30px rgba(196,30,58,.25); border: 2px solid rgba(196,30,58,.5) !important; }

/* ── History row ── */
.hist-row { display: flex; justify-content: space-between; align-items: center; padding: .8rem 1rem; border-radius: 8px; background: var(--card2); border: 1px solid var(--border); margin-bottom: .4rem; }
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# ░░  DATABASE LAYER  ░░
# In-memory for MVP. Replace with Supabase for production.
# ════════════════════════════════════════════════════════════════

PLANS = {
    "free":       {"name": "Free Trial",  "reports": 3,      "price_inr": 0,    "price_usd": 0,    "color": "#94A3B8", "team_seats": 1},
    "starter":    {"name": "Starter",     "reports": 25,     "price_inr": 499,  "price_usd": 6,    "color": "#34D399", "team_seats": 2},
    "pro":        {"name": "Pro",         "reports": 100,    "price_inr": 1299, "price_usd": 16,   "color": "#F5D76E", "team_seats": 5},
    "enterprise": {"name": "Enterprise",  "reports": 999999, "price_inr": 0,    "price_usd": 0,    "color": "#F87171", "team_seats": 25},
    "admin":      {"name": "Admin",       "reports": 999999, "price_inr": 0,    "price_usd": 0,    "color": "#FCA5A5", "team_seats": 99},
}

def _db():
    """Get or initialise the in-memory database."""
    if 'vvl_db' not in st.session_state:
        st.session_state.vvl_db = {
            "users": {
                "admin@vvl.com": {
                    "uid": "admin-001",
                    "name": "VVL Admin",
                    "password_hash": _hash("admin123"),
                    "plan": "admin",
                    "reports_used": 0,
                    "reports_limit": 999999,
                    "joined": "2024-01-01",
                    "active": True,
                    "team_id": None,
                    "report_history": [],
                },
                "demo@vvl.com": {
                    "uid": "demo-001",
                    "name": "Demo User",
                    "password_hash": _hash("demo123"),
                    "plan": "starter",
                    "reports_used": 2,
                    "reports_limit": 25,
                    "joined": datetime.now().strftime('%Y-%m-%d'),
                    "active": True,
                    "team_id": None,
                    "report_history": [
                        {"id": "r001", "filename": "sales_q1.xlsx", "domain": "Sales",    "date": "2024-12-01", "rows": 1240, "status": "complete"},
                        {"id": "r002", "filename": "hr_data.xlsx",  "domain": "HR",       "date": "2024-12-03", "rows": 430,  "status": "complete"},
                    ],
                },
            },
            "teams": {},
            "admin_api_override": "",   # Admin can override env-var key here
        }
    return st.session_state.vvl_db

def _hash(s): return hashlib.sha256(s.encode()).hexdigest()

def _user(email=None):
    email = email or st.session_state.get('vvl_email', '')
    return _db()['users'].get(email, {})

def _save_user(email, data):
    _db()['users'][email] = data

def _get_api_key():
    """Priority: admin override → Streamlit secrets → env variable."""
    override = _db().get('admin_api_override', '')
    if override and len(override) > 20:
        return override
    try:
        return st.secrets["CLAUDE_API_KEY"]
    except Exception:
        return os.environ.get("CLAUDE_API_KEY", "")

def _add_report_history(email, filename, domain, rows):
    u = _user(email)
    if 'report_history' not in u:
        u['report_history'] = []
    u['report_history'].insert(0, {
        "id": str(uuid.uuid4())[:8],
        "filename": filename,
        "domain": domain,
        "date": datetime.now().strftime('%Y-%m-%d'),
        "rows": rows,
        "status": "complete",
    })
    _db()['users'][email] = u


# ════════════════════════════════════════════════════════════════
# ░░  AUTH  ░░
# ════════════════════════════════════════════════════════════════

def auth_login(email, password):
    email = email.strip().lower()
    users = _db()['users']
    if email not in users:
        return False, "Account not found"
    u = users[email]
    if not u.get('active', True):
        return False, "Account is deactivated. Contact support."
    if u['password_hash'] != _hash(password):
        return False, "Incorrect password"
    st.session_state['vvl_logged_in'] = True
    st.session_state['vvl_email'] = email
    st.session_state['vvl_page'] = 'Dashboard'
    return True, "OK"

def auth_register(email, name, password):
    email = email.strip().lower()
    users = _db()['users']
    if not email or '@' not in email:
        return False, "Enter a valid email"
    if not name.strip():
        return False, "Enter your name"
    if len(password) < 6:
        return False, "Password must be at least 6 characters"
    if email in users:
        return False, "Email already registered"
    users[email] = {
        "uid": str(uuid.uuid4())[:8],
        "name": name.strip(),
        "password_hash": _hash(password),
        "plan": "free",
        "reports_used": 0,
        "reports_limit": PLANS["free"]["reports"],
        "joined": datetime.now().strftime('%Y-%m-%d'),
        "active": True,
        "team_id": None,
        "report_history": [],
    }
    return True, "Account created! Sign in below."

def is_admin():
    return _user().get('plan') == 'admin'

def can_run():
    u = _user()
    return u.get('reports_used', 0) < u.get('reports_limit', 0)

def do_run():
    """Increment report counter."""
    email = st.session_state.get('vvl_email', '')
    u = _db()['users'].get(email, {})
    if u:
        u['reports_used'] = u.get('reports_used', 0) + 1
        _db()['users'][email] = u


# ════════════════════════════════════════════════════════════════
# ░░  AI ENGINE  ░░
# ════════════════════════════════════════════════════════════════

def _call_ai(prompt, max_tokens=2500):
    key = _get_api_key()
    if not key:
        raise ValueError("AI service not configured. Contact admin.")
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=key)
        r = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=max_tokens,
            messages=[{"role": "user", "content": prompt}]
        )
        text = r.content[0].text.strip()
        return re.sub(r'```[a-z]*\n?', '', text).strip()
    except ImportError:
        raise ValueError("AI service temporarily unavailable.")
    except Exception as e:
        raise ValueError(str(e))

def ai_analyze(filename, summary):
    t = _call_ai(f"""Expert BI analyst. Analyze dataset. Reply ONLY valid JSON, no markdown.

File: {filename}
{summary}

Return:
{{
  "data_description": "2-3 sentence plain-English summary",
  "domain": "Sales|Finance|HR|Inventory|Marketing|Operations|Other",
  "date_column": "column name or null",
  "numeric_columns": ["list"],
  "category_columns": ["list"],
  "kpis": ["KPI 1: name — what it measures","KPI 2","KPI 3","KPI 4","KPI 5"],
  "quality_issues": ["issue 1","issue 2"],
  "insights": ["business insight 1","insight 2","insight 3"]
}}""", 2000)
    try:
        return json.loads(t)
    except Exception:
        return {"data_description": "Dataset processed.", "domain": "General",
                "date_column": None, "numeric_columns": [], "category_columns": [],
                "kpis": ["Row Count"], "quality_issues": [], "insights": ["Ready for analysis."]}

def ai_dax(filename, df, ai):
    tbl = re.sub(r'[^a-zA-Z0-9_]', '_', filename.replace('.xlsx','').replace('.csv',''))
    cols = {c: {"type": str(df[c].dtype), "sample": [str(v) for v in df[c].dropna().unique()[:3].tolist()]} for c in df.columns}
    t = _call_ai(f"""Power BI DAX expert. 10-12 production measures. ONLY valid JSON.

Table: '{tbl}'  Domain: {ai.get('domain','General')}
Columns: {json.dumps(cols)}
Date: {ai.get('date_column','none')}  Numeric: {ai.get('numeric_columns',[])}

Include: totals, averages, count, % of total, YTD/MTD (if date), MoM growth (if date), RANKX, max, min, conditional.

{{"Measure Name": "DAX = formula"}}""", 3000)
    try:
        return json.loads(t), tbl
    except Exception:
        return {f"Row Count": f"Row Count = COUNTROWS('{tbl}')"}, tbl

def ai_pq(filename, df_raw, df_clean, actions):
    ct = {c: ("type number" if 'int' in str(df_clean[c].dtype) or 'float' in str(df_clean[c].dtype)
              else "type datetime" if 'datetime' in str(df_clean[c].dtype) else "type text")
          for c in df_clean.columns}
    t = _call_ai(f"""Power Query M expert. Complete M script only — no markdown.

Source: {filename}
Types: {json.dumps(ct)}
Cleaning done: {json.dumps(actions)}

let...in script: load Excel, promote headers, remove dupes, empty rows, trim text, set types. Comment each step.""", 2500)
    if not t.strip().startswith('let'):
        p = ", ".join([f'{{"{c}", {v}}}' for c,v in ct.items()])
        return f'let\n    // Load file — update path\n    Source = Excel.Workbook(File.Contents("C:\\\\Path\\\\{filename}"), null, true),\n    Sheet = Source{{0}}[Data],\n    Headers = Table.PromoteHeaders(Sheet, [PromoteAllScalars=true]),\n    NoDupes = Table.Distinct(Headers),\n    NoEmpty = Table.SelectRows(NoDupes, each not List.IsEmpty(List.RemoveMatchingItems(Record.FieldValues(_), {{null, ""}}))),\n    Trimmed = Table.TransformColumns(NoEmpty, {{}}, Text.Trim),\n    Typed = Table.TransformColumnTypes(Trimmed, {{{p}}})\nin\n    Typed'
    return t


# ════════════════════════════════════════════════════════════════
# ░░  DATA CLEANING  ░░
# ════════════════════════════════════════════════════════════════

def _sum(df):
    lines = [f"Shape: {df.shape[0]} rows x {df.shape[1]} cols\nColumns:"]
    for c in df.columns:
        dtype, nulls, uniq = str(df[c].dtype), int(df[c].isnull().sum()), int(df[c].nunique())
        info = f"min={df[c].min():.1f},max={df[c].max():.1f},mean={df[c].mean():.1f}" if dtype in ['float64','int64','float32','int32'] else f"samples={str(df[c].dropna().unique()[:4].tolist())}"
        lines.append(f"  {c}|{dtype}|{nulls} nulls|{uniq} unique|{info}")
    return "\n".join(lines)

def clean(df):
    d = df.copy(); acts = []
    old = list(d.columns)
    d.columns = [re.sub(r'\s+','_',re.sub(r'[^a-zA-Z0-9\s_]','',str(c)).strip()).strip('_') or 'Col' for c in d.columns]
    ren = [(o,n) for o,n in zip(old,d.columns) if o!=n]
    if ren: acts.append(f"Renamed {len(ren)} column(s) — removed special chars")
    b = len(d); d.dropna(how='all',inplace=True); d.dropna(axis=1,how='all',inplace=True)
    if b-len(d)>0: acts.append(f"Removed {b-len(d)} fully empty rows")
    dp = d.duplicated().sum()
    if dp>0: d.drop_duplicates(inplace=True); acts.append(f"Removed {dp} duplicate rows")
    for c in d.select_dtypes(include='object').columns:
        d[c] = d[c].apply(lambda x: x.strip() if isinstance(x,str) else x).replace('',np.nan)
    acts.append("Trimmed whitespace from text columns")
    cn=[]
    for c in d.select_dtypes(include='object').columns:
        try:
            tmp = pd.to_numeric(d[c].astype(str).str.replace(',','').str.replace('$','').str.replace('%','').str.strip(),errors='coerce')
            if tmp.notna().mean()>0.8: d[c]=tmp; cn.append(c)
        except: pass
    if cn: acts.append(f"Converted to numeric: {', '.join(cn)}")
    cd=[]
    for c in d.select_dtypes(include='object').columns:
        if any(k in c.lower() for k in ['date','time','year','month','day','dt','created','updated']):
            try:
                r2=pd.to_datetime(d[c],infer_datetime_format=True,errors='coerce')
                if r2.notna().mean()>0.7: d[c]=r2; cd.append(c)
            except: pass
    if cd: acts.append(f"Parsed date columns: {', '.join(cd)}")
    fl=[]
    for c in d.select_dtypes(include=[np.number]).columns:
        n=d[c].isnull().sum()
        if n>0: d[c]=d[c].fillna(d[c].median()); fl.append(f"{c}({n})")
    if fl: acts.append(f"Filled nulls with median: {', '.join(fl)}")
    d.reset_index(drop=True,inplace=True)
    return d, acts


# ════════════════════════════════════════════════════════════════
# ░░  CHARTS  ░░
# ════════════════════════════════════════════════════════════════

PAL = ['#C41E3A','#D4A017','#F5D76E','#8B0000','#DC143C','#B8860B','#E8A0A0','#F0D080']
LO = dict(paper_bgcolor='#0F1420', plot_bgcolor='#0F1420',
          font=dict(color='#E8E2D9', family='Outfit'),
          xaxis=dict(gridcolor='rgba(196,30,58,.07)', zerolinecolor='rgba(196,30,58,.15)'),
          yaxis=dict(gridcolor='rgba(196,30,58,.07)', zerolinecolor='rgba(196,30,58,.15)'),
          margin=dict(t=50, l=20, r=20, b=30),
          hoverlabel=dict(bgcolor='#141B2D', bordercolor='#C41E3A', font_color='#E8E2D9'),
          legend=dict(bgcolor='rgba(0,0,0,0)', font_color='#94A3B8'))

def charts(df, ai):
    try:
        import plotly.graph_objects as go
    except Exception:
        return []
    c = []
    nm = [x for x in (ai.get('numeric_columns') or []) if x in df.columns] or list(df.select_dtypes(include=[np.number]).columns)
    cm = [x for x in (ai.get('category_columns') or []) if x in df.columns] or [x for x in df.select_dtypes(include='object').columns if df[x].nunique() < 30]
    dc = ai.get('date_column'); dc = dc if dc and dc in df.columns else None

    if dc and nm:
        try:
            cols = nm[:2]; dt = df[[dc]+cols].dropna().sort_values(dc)
            fig = go.Figure()
            for i,col in enumerate(cols):
                fig.add_trace(go.Scatter(x=dt[dc],y=dt[col],name=col,mode='lines+markers',
                    line=dict(color=PAL[i],width=2.5,shape='spline'),marker=dict(size=4),
                    fill='tozeroy' if i==0 else 'none',fillcolor='rgba(196,30,58,0.04)' if i==0 else None))
            fig.update_layout(**LO, title="Trend Over Time", hovermode='x unified')
            c.append(fig)
        except Exception: pass

    if cm and nm:
        try:
            ca,nu = cm[0],nm[0]
            dfb = df.groupby(ca)[nu].sum().reset_index().sort_values(nu,ascending=False).head(15)
            fig = go.Figure(go.Bar(x=dfb[ca],y=dfb[nu],
                marker=dict(color=dfb[nu],colorscale=[[0,'#4A0010'],[.5,'#C41E3A'],[1,'#F5D76E']]),
                hovertemplate=f'<b>%{{x}}</b><br>{nu}: %{{y:,.0f}}<extra></extra>'))
            fig.update_layout(**LO, title=f"{nu} by {ca}")
            c.append(fig)
        except Exception: pass

    if cm and nm and df[cm[0]].nunique()<=12:
        try:
            ca,nu=cm[0],nm[0]; dfp=df.groupby(ca)[nu].sum().reset_index()
            fig=go.Figure(go.Pie(labels=dfp[ca],values=dfp[nu],hole=0.48,
                marker=dict(colors=PAL,line=dict(color='#07090F',width=2)),
                hovertemplate='<b>%{label}</b><br>%{value:,.0f}  (%{percent})<extra></extra>'))
            fig.update_layout(**LO, title=f"{nu} Share")
            c.append(fig)
        except Exception: pass

    if nm:
        try:
            col=nm[0]
            fig=go.Figure(go.Histogram(x=df[col],nbinsx=30,
                marker=dict(color='#C41E3A',opacity=.8,line=dict(color='#07090F',width=1))))
            fig.update_layout(**LO, title=f"Distribution of {col}")
            c.append(fig)
        except Exception: pass

    if len(nm)>=3:
        try:
            corr=df[nm[:8]].corr()
            fig=go.Figure(go.Heatmap(z=corr.values,x=corr.columns,y=corr.columns,
                colorscale=[[0,'#0F1420'],[.5,'#8B0000'],[1,'#F5D76E']],
                zmin=-1,zmax=1,text=corr.values.round(2),texttemplate='%{text}',textfont=dict(size=10)))
            fig.update_layout(**LO, title="Correlation Matrix")
            c.append(fig)
        except Exception: pass
    return c


# ════════════════════════════════════════════════════════════════
# ░░  EXPORT BUILDERS  ░░
# ════════════════════════════════════════════════════════════════

def build_excel(df_c, ai, dax, acts, fname):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb=Workbook(); CR='C41E3A'; GO='D4A017'; G2='F5D76E'; DK='07090F'; CA='0F1420'; W='FFFFFF'; AL='141B2D'
    def H(cell,bg=None): cell.font=Font(bold=True,color=W,size=10,name='Calibri'); cell.fill=PatternFill('solid',start_color=bg or DK); cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    def D(cell,row): cell.font=Font(name='Calibri',size=10,color='E8E2D9'); cell.fill=PatternFill('solid',start_color=CA if row%2 else AL); cell.alignment=Alignment(vertical='center')

    ws=wb.active; ws.title="Cleaned Data"
    ws.merge_cells(f'A1:{get_column_letter(len(df_c.columns))}1')
    ws['A1']=f"Vampire Venom Lab  ·  {df_c.shape[0]:,} rows × {df_c.shape[1]} cols  ·  {fname}"
    ws['A1'].font=Font(bold=True,size=11,color=G2,name='Calibri'); ws['A1'].alignment=Alignment(horizontal='center',vertical='center'); ws['A1'].fill=PatternFill('solid',start_color=DK); ws.row_dimensions[1].height=25
    for ci,col in enumerate(df_c.columns,1): H(ws.cell(row=2,column=ci,value=col),bg=CA); ws.row_dimensions[2].height=20
    for ri,row in enumerate(df_c.itertuples(index=False),3):
        for ci,val in enumerate(row,1):
            if hasattr(val,'item'): val=val.item()
            elif not isinstance(val,str):
                try:
                    if pd.isna(val): val=None
                except: pass
            D(ws.cell(row=ri,column=ci,value=val),ri)
    for ci,col in enumerate(df_c.columns,1):
        mx=max(len(str(col)),df_c.iloc[:,ci-1].astype(str).str.len().max() if len(df_c) else 0)
        ws.column_dimensions[get_column_letter(ci)].width=min(max(mx+2,12),40)
    tr=len(df_c)+3; ws.cell(row=tr,column=1,value="TOTALS").font=Font(bold=True,color=G2,name='Calibri')
    for ci,col in enumerate(df_c.columns,1):
        if str(df_c[col].dtype) in ['float64','int64','float32','int32']:
            cl=get_column_letter(ci); c=ws.cell(row=tr,column=ci,value=f'=SUM({cl}3:{cl}{tr-1})'); c.font=Font(bold=True,color=G2,name='Calibri'); c.fill=PatternFill('solid',start_color=DK)
    ws.freeze_panes='A3'

    ws3=wb.create_sheet("DAX Formulas"); ws3.column_dimensions['A'].width=35; ws3.column_dimensions['B'].width=90
    ws3.merge_cells('A1:B1'); ws3['A1']="DAX Formulas — Modeling → New Measure → Paste"
    ws3['A1'].font=Font(bold=True,size=11,color=G2,name='Calibri'); ws3['A1'].alignment=Alignment(horizontal='center'); ws3['A1'].fill=PatternFill('solid',start_color=DK); ws3.row_dimensions[1].height=25
    for ci,h in enumerate(['Measure','Formula'],1): H(ws3.cell(row=2,column=ci),bg=CA); ws3.cell(row=2,column=ci).value=h
    for ri,(n,f) in enumerate(dax.items(),3):
        nc=ws3.cell(row=ri,column=1,value=n); nc.font=Font(bold=True,name='Calibri',size=10,color=G2); nc.fill=PatternFill('solid',start_color=AL if ri%2 else CA)
        fc=ws3.cell(row=ri,column=2,value=f); fc.font=Font(name='Courier New',size=9,color='F5D76E'); fc.alignment=Alignment(wrap_text=True); fc.fill=PatternFill('solid',start_color='08100A' if ri%2 else '06090F'); ws3.row_dimensions[ri].height=42

    ws4=wb.create_sheet("Summary"); ws4.column_dimensions['A'].width=38; ws4.column_dimensions['B'].width=55
    ws4.merge_cells('A1:B1'); ws4['A1']="Vampire Venom Lab — Analysis Summary"
    ws4['A1'].font=Font(bold=True,size=13,color=G2,name='Calibri'); ws4['A1'].alignment=Alignment(horizontal='center'); ws4['A1'].fill=PatternFill('solid',start_color=DK); ws4.row_dimensions[1].height=28
    r=3
    def sec(ws,t,row):
        ws.merge_cells(f'A{row}:B{row}'); c=ws.cell(row=row,column=1,value=f"  {t}")
        c.font=Font(bold=True,color=W,size=11,name='Calibri'); c.fill=PatternFill('solid',start_color=CA); c.alignment=Alignment(horizontal='left',vertical='center'); ws.row_dimensions[row].height=20; return row+1
    def lst(ws,items,row):
        for item in items:
            ws.merge_cells(f'A{row}:B{row}'); c=ws.cell(row=row,column=1,value=f"  • {item}")
            c.font=Font(name='Calibri',size=10,color='94A3B8'); c.fill=PatternFill('solid',start_color=AL if row%2 else CA); c.alignment=Alignment(wrap_text=True); ws.row_dimensions[row].height=20; row+=1
        return row+1
    r=sec(ws4,"📋 OVERVIEW",r)
    for lb,vl in [("File",fname),("Domain",ai.get('domain','')),("Rows",f"{df_c.shape[0]:,}"),("Cols",str(df_c.shape[1])),("Generated",datetime.now().strftime('%Y-%m-%d %H:%M'))]:
        ws4.cell(row=r,column=1,value=lb).font=Font(bold=True,name='Calibri',size=10,color='94A3B8'); ws4.cell(row=r,column=2,value=str(vl)).font=Font(name='Calibri',size=10,color='E8E2D9')
        for ci in [1,2]: ws4.cell(row=r,column=ci).fill=PatternFill('solid',start_color=AL if r%2 else CA); r+=1
    r+=1; r=sec(ws4,"📝 DESCRIPTION",r)
    ws4.merge_cells(f'A{r}:B{r}'); c=ws4.cell(row=r,column=1,value=ai.get('data_description','')); c.font=Font(name='Calibri',size=10,color='B0A899'); c.alignment=Alignment(wrap_text=True); c.fill=PatternFill('solid',start_color=AL); ws4.row_dimensions[r].height=55; r+=2
    r=sec(ws4,"🔧 CLEANING",r); r=lst(ws4,acts,r)
    r=sec(ws4,"🎯 KPIs",r); r=lst(ws4,ai.get('kpis',[]),r)
    r=sec(ws4,"💡 INSIGHTS",r); r=lst(ws4,ai.get('insights',[]),r)

    ws5=wb.create_sheet("Data Dictionary")
    for ci,w in enumerate([30,15,16,14,16,28],1): ws5.column_dimensions[get_column_letter(ci)].width=w
    ws5.merge_cells('A1:F1'); ws5['A1']="Data Dictionary"
    ws5['A1'].font=Font(bold=True,size=12,color=G2,name='Calibri'); ws5['A1'].alignment=Alignment(horizontal='center'); ws5['A1'].fill=PatternFill('solid',start_color=DK)
    for ci,h in enumerate(['Column','Type','Non-Null','Nulls','Unique','Sample Values'],1): H(ws5.cell(row=2,column=ci),bg=CA); ws5.cell(row=2,column=ci).value=h
    for ri,col in enumerate(df_c.columns,3):
        sv=', '.join(str(v) for v in df_c[col].dropna().unique()[:3])
        for ci,val in enumerate([col,str(df_c[col].dtype),int(df_c[col].count()),int(df_c[col].isnull().sum()),int(df_c[col].nunique()),sv],1):
            c=ws5.cell(row=ri,column=ci,value=val); c.font=Font(name='Calibri',size=10,color='E8E2D9'); c.fill=PatternFill('solid',start_color=AL if ri%2 else CA)

    out=io.BytesIO(); wb.save(out); return out.getvalue()

def build_pdf(df_c, ai, dax, acts, fname):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,rightMargin=2*cm,leftMargin=2*cm,topMargin=2*cm,bottomMargin=2*cm)
    st2=getSampleStyleSheet()
    CC=colors.HexColor('#C41E3A'); CG=colors.HexColor('#D4A017'); CG2=colors.HexColor('#F5D76E')
    CD=colors.HexColor('#07090F'); CK=colors.HexColor('#0F1420'); CT=colors.HexColor('#E8E2D9'); CM=colors.HexColor('#5A6070')
    H1=ParagraphStyle('H1',parent=st2['Heading1'],textColor=CG2,fontSize=22,spaceAfter=4,alignment=TA_CENTER,fontName='Helvetica-Bold')
    H2=ParagraphStyle('H2',parent=st2['Heading2'],textColor=CT,fontSize=13,spaceBefore=16,spaceAfter=6,fontName='Helvetica-Bold')
    BO=ParagraphStyle('B',parent=st2['Normal'],fontSize=10,leading=17,spaceAfter=4,textColor=CM)
    CO=ParagraphStyle('C',parent=st2['Normal'],fontSize=8,fontName='Courier',leading=13,textColor=CG2,backColor=colors.HexColor('#06090F'),leftIndent=8,rightIndent=8,spaceAfter=4)
    SU=ParagraphStyle('S',parent=st2['Normal'],fontSize=10,textColor=CM,alignment=TA_CENTER,spaceAfter=14)
    story=[]
    story.append(Spacer(1,.5*cm)); story.append(Paragraph("🧪 Vampire Venom Lab",H1))
    story.append(Paragraph("Business Intelligence Report",ParagraphStyle('T',parent=st2['Normal'],fontSize=14,textColor=CG,alignment=TA_CENTER,spaceAfter=5,fontName='Helvetica-Bold')))
    story.append(Paragraph(f"{datetime.now().strftime('%B %d, %Y  ·  %H:%M')}  ·  {fname}  ·  {df_c.shape[0]:,} rows  ·  {ai.get('domain','General')} domain",SU))
    story.append(HRFlowable(width='100%',color=CC,thickness=1.5)); story.append(Spacer(1,.3*cm))
    story.append(Paragraph("Data Overview",H2)); story.append(Paragraph(ai.get('data_description',''),BO))
    td=[['Metric','Value']]
    for l,v in [("File",fname),("Domain",ai.get('domain','')),("Rows",f"{df_c.shape[0]:,}"),("Columns",str(df_c.shape[1])),("Missing (after clean)",str(df_c.isnull().sum().sum()))]: td.append([l,v])
    t=Table(td,colWidths=[8*cm,8*cm],repeatRows=1)
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),CC),('TEXTCOLOR',(0,0),(-1,0),CG2),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),('ROWBACKGROUNDS',(0,1),(-1,-1),[CK,colors.HexColor('#141B2D')]),('TEXTCOLOR',(0,1),(-1,-1),CT),('GRID',(0,0),(-1,-1),.3,colors.HexColor('#1e2535')),('ALIGN',(1,1),(1,-1),'RIGHT'),('PADDING',(0,0),(-1,-1),6)])); story.append(t)
    def add(title,items): story.append(Paragraph(title,H2)); [story.append(Paragraph(f"• {i}",BO)) for i in items]
    add("Cleaning Actions",acts); add("Recommended KPIs",ai.get('kpis',[])); add("Business Insights",ai.get('insights',[]))
    nd=df_c.select_dtypes(include=[np.number])
    if not nd.empty:
        story.append(Paragraph("Numeric Statistics",H2))
        sd=[['Column','Sum','Average','Min','Max']]
        for col in nd.columns: sd.append([col,f"{nd[col].sum():,.1f}",f"{nd[col].mean():,.1f}",f"{nd[col].min():,.1f}",f"{nd[col].max():,.1f}"])
        t2=Table(sd,repeatRows=1); t2.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),CG),('TEXTCOLOR',(0,0),(-1,0),CD),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),('ROWBACKGROUNDS',(0,1),(-1,-1),[CK,colors.HexColor('#141B2D')]),('TEXTCOLOR',(0,1),(-1,-1),CT),('GRID',(0,0),(-1,-1),.3,colors.HexColor('#1e2535')),('ALIGN',(1,1),(-1,-1),'RIGHT'),('PADDING',(0,0),(-1,-1),6)])); story.append(t2)
    if dax:
        story.append(Paragraph("DAX Formulas",H2))
        story.append(Paragraph("Power BI → Modeling → New Measure → Paste formula below",ParagraphStyle('n',parent=st2['Normal'],fontSize=9,textColor=CM,spaceAfter=8)))
        for name,formula in dax.items():
            story.append(Paragraph(f"<b><font color='#F5D76E'>{name}</font></b>",ParagraphStyle('mn',parent=st2['Normal'],fontSize=10,textColor=CG2,spaceAfter=3,spaceBefore=10)))
            story.append(Paragraph(textwrap.fill(formula,90).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;'),CO))
    doc.build(story); return buf.getvalue()


# ════════════════════════════════════════════════════════════════
# ░░  PAGE: LOGIN / REGISTER  ░░
# ════════════════════════════════════════════════════════════════

def page_auth():
    # Hero
    st.markdown("""
    <div style="text-align:center;padding:2.5rem 0 1.5rem 0;">
        <div style="font-family:'Cinzel Decorative',serif;font-size:2.8rem;font-weight:700;
                    background:linear-gradient(135deg,#F5D76E 0%,#C41E3A 55%,#D4A017 100%);
                    -webkit-background-clip:text;-webkit-text-fill-color:transparent;
                    line-height:1.1;margin-bottom:.4rem;">
            🧪 Vampire Venom Lab
        </div>
        <div style="font-family:'Cinzel',serif;font-size:.85rem;color:#5A6070;letter-spacing:3px;
                    text-transform:uppercase;margin-bottom:.8rem;">
            Enterprise BI Intelligence Platform
        </div>
        <div style="font-size:1rem;color:#3D4455;max-width:480px;margin:0 auto;line-height:1.6;">
            Transform raw Excel data into Power BI-ready reports,<br>
            DAX formulas and business intelligence — in seconds.
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<hr>", unsafe_allow_html=True)

    _, col, _ = st.columns([1, 1.4, 1])
    with col:
        t_login, t_reg = st.tabs(["🔐 Sign In", "📝 Create Account"])

        with t_login:
            st.markdown("<br>", unsafe_allow_html=True)
            em = st.text_input("Email address", placeholder="you@company.com", key="l_em")
            pw = st.text_input("Password", type="password", placeholder="••••••••", key="l_pw")
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Sign In →", use_container_width=True, key="l_btn"):
                if em and pw:
                    ok, msg = auth_login(em, pw)
                    if ok:
                        st.success("Welcome back!")
                        time.sleep(.4)
                        st.rerun()
                    else:
                        st.error(f"✗  {msg}")
                else:
                    st.warning("Please enter email and password")
            st.markdown("""
            <div style="text-align:center;margin-top:1.2rem;font-size:.8rem;color:#374151;">
                Demo account — <code style="color:#D4A017;">demo@vvl.com</code> &nbsp;/&nbsp; <code style="color:#D4A017;">demo123</code>
            </div>""", unsafe_allow_html=True)

        with t_reg:
            st.markdown("<br>", unsafe_allow_html=True)
            rn = st.text_input("Full Name", placeholder="Your Name", key="r_n")
            re_ = st.text_input("Email address", placeholder="you@company.com", key="r_e")
            rp = st.text_input("Password", type="password", placeholder="Min. 6 characters", key="r_p")
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Create Free Account →", use_container_width=True, key="r_btn"):
                ok, msg = auth_register(re_, rn, rp)
                if ok:
                    st.success(f"✓  {msg}")
                else:
                    st.error(f"✗  {msg}")

    # Pricing strip
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown('<div style="text-align:center;font-family:\'Cinzel\',serif;font-size:.72rem;color:#D4A017;letter-spacing:3px;text-transform:uppercase;margin-bottom:1.4rem;">Subscription Plans</div>', unsafe_allow_html=True)
    cols = st.columns(4)
    for col, pk in zip(cols, ['free','starter','pro','enterprise']):
        p = PLANS[pk]
        highlight = pk == 'pro'
        border = "border:2px solid rgba(196,30,58,.5);box-shadow:0 0 28px rgba(196,30,58,.18);" if highlight else ""
        badge = '<div style="font-size:.68rem;color:#C41E3A;font-weight:700;letter-spacing:1px;margin-bottom:.4rem;">🔥 POPULAR</div>' if highlight else '<div style="height:1.1rem;"></div>'
        with col:
            st.markdown(f"""
            <div class="vvl-card" style="{border}text-align:center;padding:1.6rem 1rem;">
                {badge}
                <div style="font-family:'Cinzel',serif;font-size:.95rem;font-weight:600;color:#E8E2D9;margin-bottom:.3rem;">{p['name']}</div>
                <div style="font-size:1.7rem;font-weight:700;color:{p['color']};margin:.5rem 0;">
                    {"Free" if p['price_inr']==0 and pk!='enterprise' else ("Custom" if pk=='enterprise' else f"₹{p['price_inr']:,}")}
                    <span style="font-size:.75rem;color:#5A6070;">{"/mo" if pk not in ["free","enterprise"] else ""}</span>
                </div>
                <div style="font-size:.82rem;color:#5A6070;">{p['reports'] if p['reports']<999999 else "Unlimited"} reports/month</div>
                <div style="font-size:.78rem;color:#3D4455;margin-top:.3rem;">{p['team_seats']} team seat{"s" if p['team_seats']>1 else ""}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div style="text-align:center;font-size:.75rem;color:#1A2030;">© 2024 Vampire Venom Lab · All rights reserved</div>', unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# ░░  SIDEBAR  ░░
# ════════════════════════════════════════════════════════════════

def render_sidebar():
    u = _user(); plan = u.get('plan','free'); email = st.session_state.get('vvl_email','')
    used = u.get('reports_used',0); lim = u.get('reports_limit',3)
    pct = min(int(used/max(lim,1)*100), 100)
    bar_color = "#EF4444" if pct>80 else "#F59E0B" if pct>55 else "#10B981"

    with st.sidebar:
        # Brand
        st.markdown("""
        <div style="padding:1.2rem 0 .8rem 0;text-align:center;">
            <div style="font-family:'Cinzel',serif;font-size:1.25rem;font-weight:700;
                        background:linear-gradient(135deg,#F5D76E,#C41E3A);
                        -webkit-background-clip:text;-webkit-text-fill-color:transparent;">
                🧪 Vampire Venom Lab
            </div>
            <div style="font-size:.65rem;color:#2A3040;letter-spacing:2px;text-transform:uppercase;margin-top:2px;">
                BI Intelligence Platform
            </div>
        </div>""", unsafe_allow_html=True)

        st.markdown("<hr>", unsafe_allow_html=True)

        # User card
        st.markdown(f"""
        <div class="vvl-card" style="margin-bottom:.8rem;padding:1rem;">
            <div style="font-size:.75rem;color:#5A6070;margin-bottom:3px;">Signed in as</div>
            <div style="font-weight:600;font-size:.92rem;color:#E8E2D9;margin-bottom:4px;">{u.get('name','User')}</div>
            <div style="font-size:.75rem;color:#3D4455;margin-bottom:7px;">{email}</div>
            <span class="plan-badge badge-{plan}">{PLANS.get(plan,{}).get('name',plan)}</span>
        </div>""", unsafe_allow_html=True)

        # Usage meter
        if plan not in ['admin','enterprise']:
            st.markdown(f"""
            <div style="margin-bottom:1rem;">
                <div style="display:flex;justify-content:space-between;margin-bottom:4px;">
                    <span style="font-size:.74rem;color:#5A6070;">Reports</span>
                    <span style="font-size:.74rem;color:#E8E2D9;font-weight:600;">{used} / {lim if lim<999999 else "∞"}</span>
                </div>
                <div class="usage-bar-bg">
                    <div class="usage-bar-fill" style="background:{bar_color};width:{pct}%;"></div>
                </div>
                {f'<div style="font-size:.7rem;color:#EF4444;margin-top:3px;">Limit reached — upgrade to continue</div>' if pct>=100 else ''}
            </div>""", unsafe_allow_html=True)

        st.markdown("<hr>", unsafe_allow_html=True)

        # Nav
        st.markdown('<span class="vvl-label">Navigation</span>', unsafe_allow_html=True)
        current = st.session_state.get('vvl_page','Dashboard')
        nav_items = [
            ("🏠", "Dashboard"),
            ("🧪", "BI Analyzer"),
            ("📋", "Report History"),
            ("📊", "Usage & Billing"),
            ("👥", "Team Access"),
            ("💎", "Upgrade Plan"),
        ]
        if plan == 'admin':
            nav_items.append(("⚙️", "Admin Panel"))

        for icon, page in nav_items:
            active = current == page
            bg = "background:rgba(196,30,58,.15);border:1px solid rgba(196,30,58,.3);" if active else "border:1px solid transparent;"
            tc = "color:#F5D76E;" if active else "color:#5A6070;"
            fw = "font-weight:700;" if active else "font-weight:400;"
            st.markdown(f'<div style="padding:.55rem .8rem;border-radius:8px;{bg}display:flex;align-items:center;gap:9px;margin-bottom:3px;cursor:pointer;"><span style="font-size:1rem;">{icon}</span><span style="font-size:.88rem;{tc}{fw}">{page}</span></div>', unsafe_allow_html=True)
            if st.button(f"{icon} {page}", key=f"nav_{page}", use_container_width=True):
                st.session_state['vvl_page'] = page
                st.rerun()

        st.markdown("<hr>", unsafe_allow_html=True)

        # Pipeline steps
        st.markdown('<span class="vvl-label">Pipeline</span>', unsafe_allow_html=True)
        for n,l in [("01","Upload Excel / CSV"),("02","AI Analyzes"),("03","Auto-Clean"),("04","Generate DAX"),("05","Power Query"),("06","Build Charts"),("07","Export All")]:
            st.markdown(f'<div style="display:flex;gap:9px;margin-bottom:6px;"><span style="font-family:\'Cinzel\',serif;font-size:.62rem;color:#C41E3A;opacity:.55;min-width:20px;padding-top:1px;">{n}</span><span style="font-size:.8rem;color:#3D4455;">{l}</span></div>', unsafe_allow_html=True)

        st.markdown("<hr>", unsafe_allow_html=True)

        if st.button("⬡  Sign Out", use_container_width=True, key="signout"):
            for k in ['vvl_logged_in','vvl_email','vvl_page']:
                st.session_state.pop(k, None)
            st.rerun()

        st.markdown('<div style="font-size:.65rem;color:#0E1420;text-align:center;margin-top:.5rem;">© 2024 Vampire Venom Lab</div>', unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# ░░  PAGE: DASHBOARD  ░░
# ════════════════════════════════════════════════════════════════

def page_dashboard():
    u = _user(); plan = u.get('plan','free')
    used = u.get('reports_used',0); lim = u.get('reports_limit',3)
    history = u.get('report_history',[])

    st.markdown(f"""
    <div style="margin-bottom:1.5rem;">
        <h1 style="font-family:'Cinzel',serif;font-size:1.9rem;font-weight:700;
                   background:linear-gradient(90deg,#F5D76E,#C41E3A);
                   -webkit-background-clip:text;-webkit-text-fill-color:transparent;margin:0 0 .3rem 0;">
            Welcome back, {u.get('name','').split()[0]} 👋
        </h1>
        <div style="color:#5A6070;font-size:.9rem;">
            Here's your account overview and recent activity.
        </div>
    </div>""", unsafe_allow_html=True)

    # KPI strip
    m1,m2,m3,m4 = st.columns(4)
    m1.metric("Plan", PLANS.get(plan,{}).get('name',plan))
    m2.metric("Reports Used", used, delta=None)
    m3.metric("Reports Remaining", max(lim-used,0) if lim<999999 else "∞")
    m4.metric("Team Members", len(_db().get('teams',{}).get(u.get('team_id',''),{}).get('members',[])) or 1)

    st.markdown("<br>", unsafe_allow_html=True)

    left, right = st.columns([1.6, 1])

    with left:
        st.markdown('<span class="vvl-label">Recent Reports</span>', unsafe_allow_html=True)
        if history:
            for h in history[:5]:
                dom_color = {"Sales":"#10B981","Finance":"#F5D76E","HR":"#60A5FA","Inventory":"#F59E0B"}.get(h.get('domain',''),"#C41E3A")
                st.markdown(f"""
                <div class="hist-row">
                    <div style="display:flex;align-items:center;gap:10px;">
                        <div style="width:8px;height:8px;border-radius:50%;background:{dom_color};flex-shrink:0;"></div>
                        <div>
                            <div style="font-weight:600;font-size:.88rem;color:#E8E2D9;">{h.get('filename','')}</div>
                            <div style="font-size:.75rem;color:#5A6070;">{h.get('date','')} · {h.get('rows',0):,} rows</div>
                        </div>
                    </div>
                    <div>
                        <span style="font-size:.72rem;color:{dom_color};background:rgba(255,255,255,.05);padding:2px 10px;border-radius:99px;border:1px solid rgba(255,255,255,.08);">{h.get('domain','')}</span>
                    </div>
                </div>""", unsafe_allow_html=True)
        else:
            st.markdown('<div class="vvl-card" style="text-align:center;color:#3D4455;padding:2rem;">No reports yet. Run your first analysis in the BI Analyzer.</div>', unsafe_allow_html=True)

        if st.button("➡  Go to BI Analyzer", key="dash_go_analyzer"):
            st.session_state['vvl_page'] = 'BI Analyzer'; st.rerun()

    with right:
        st.markdown('<span class="vvl-label">Usage This Month</span>', unsafe_allow_html=True)
        pct = min(int(used/max(lim,1)*100),100) if lim<999999 else 0
        bar_color = "#EF4444" if pct>80 else "#F59E0B" if pct>55 else "#10B981"
        st.markdown(f"""
        <div class="vvl-card vvl-card-crimson" style="margin-bottom:.8rem;">
            <div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:8px;">
                <span style="font-size:2rem;font-weight:700;color:{bar_color};">{used}</span>
                <span style="font-size:.85rem;color:#5A6070;">of {lim if lim<999999 else "∞"} reports</span>
            </div>
            <div class="usage-bar-bg"><div class="usage-bar-fill" style="background:{bar_color};width:{pct}%;"></div></div>
            <div style="font-size:.75rem;color:#5A6070;margin-top:6px;">{pct}% used this billing cycle</div>
        </div>""", unsafe_allow_html=True)

        st.markdown('<span class="vvl-label">Account Details</span>', unsafe_allow_html=True)
        for l,v in [("Plan", PLANS.get(plan,{}).get('name',plan)), ("Member since", u.get('joined','')), ("Team seats", PLANS.get(plan,{}).get('team_seats',1))]:
            st.markdown(f'<div style="display:flex;justify-content:space-between;padding:.4rem 0;border-bottom:1px solid var(--border);"><span style="font-size:.82rem;color:#5A6070;">{l}</span><span style="font-size:.82rem;color:#E8E2D9;font-weight:500;">{v}</span></div>', unsafe_allow_html=True)

        if plan == 'free':
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<div class="vvl-card" style="background:linear-gradient(135deg,rgba(196,30,58,.12),rgba(212,160,23,.08));border-color:rgba(196,30,58,.3);text-align:center;padding:1rem;">', unsafe_allow_html=True)
            st.markdown('<div style="font-family:\'Cinzel\',serif;font-size:.85rem;color:#F5D76E;margin-bottom:.4rem;">Upgrade for more reports</div>', unsafe_allow_html=True)
            st.markdown('<div style="font-size:.8rem;color:#5A6070;margin-bottom:.8rem;">Starter plan: 25 reports/month for ₹499</div>', unsafe_allow_html=True)
            if st.button("💎 Upgrade Now", key="dash_upgrade"):
                st.session_state['vvl_page'] = 'Upgrade Plan'; st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# ░░  PAGE: BI ANALYZER  ░░
# ════════════════════════════════════════════════════════════════

def page_analyzer():
    u = _user(); plan = u.get('plan','free'); email = st.session_state.get('vvl_email','')

    st.markdown("""
    <h1 style="font-family:'Cinzel',serif;font-size:1.8rem;font-weight:700;
               background:linear-gradient(90deg,#F5D76E,#C41E3A);
               -webkit-background-clip:text;-webkit-text-fill-color:transparent;margin:0 0 .4rem 0;">
        🧪 BI Analyzer
    </h1>
    <div style="color:#5A6070;font-size:.9rem;margin-bottom:1.5rem;">
        Upload your Excel or CSV — AI does the full BI workflow
    </div>""", unsafe_allow_html=True)

    # Gate: limit reached
    if not can_run() and plan not in ['admin','enterprise']:
        st.markdown(f"""<div class="vvl-card" style="text-align:center;padding:2.5rem;border-color:rgba(196,30,58,.4);">
        <div style="font-size:1.8rem;margin-bottom:.8rem;">🔒</div>
        <div style="font-family:'Cinzel',serif;font-size:1.1rem;color:#F5D76E;margin-bottom:.5rem;">Report Limit Reached</div>
        <div style="color:#5A6070;font-size:.9rem;">You've used all {u.get('reports_limit',3)} reports on your {PLANS.get(plan,{}).get('name','')} plan.</div>
        </div>""", unsafe_allow_html=True)
        if st.button("💎 Upgrade Plan", key="gate_upgrade"):
            st.session_state['vvl_page'] = 'Upgrade Plan'; st.rerun()
        return

    # Gate: no API key
    if not _get_api_key():
        st.error("⚙️ AI service not configured. Administrator needs to set the API key.")
        return

    # Upload
    st.markdown('<span class="vvl-label">Upload Your Data</span>', unsafe_allow_html=True)
    uploaded = st.file_uploader("Drop Excel / CSV files", type=["xlsx","xls","csv"], accept_multiple_files=True, label_visibility="collapsed")
    if not uploaded:
        st.markdown('<div class="vvl-card" style="text-align:center;padding:2rem;color:#3D4455;"><div style="font-size:1.4rem;margin-bottom:.5rem;">📂</div>Drop Excel (.xlsx .xls) or CSV files above</div>', unsafe_allow_html=True)
        return

    # Load
    all_dfs = {}
    for uf in uploaded:
        try:
            raw = uf.read()
            df = pd.read_csv(io.BytesIO(raw)) if uf.name.endswith('.csv') else list(pd.read_excel(io.BytesIO(raw),sheet_name=None).values())[0]
            all_dfs[uf.name] = df
        except Exception as e:
            st.error(f"Cannot read **{uf.name}**: {e}")
    if not all_dfs: return

    # Preview
    st.markdown("<br>", unsafe_allow_html=True)
    for fname, df in all_dfs.items():
        with st.expander(f"📄 {fname}  —  {df.shape[0]:,} rows × {df.shape[1]} cols", expanded=True):
            m1,m2,m3,m4 = st.columns(4)
            m1.metric("Rows",f"{df.shape[0]:,}"); m2.metric("Columns",df.shape[1])
            m3.metric("Missing",f"{df.isnull().sum().sum():,}"); m4.metric("Duplicates",f"{df.duplicated().sum():,}")
            st.dataframe(df.head(8), use_container_width=True)

    st.markdown("<br>", unsafe_allow_html=True)
    _, cc, _ = st.columns([1,2,1])
    with cc:
        run = st.button("🧪  Run Full AI Analysis", use_container_width=True, key="run_analysis")
    if not run: return

    # Count usage & save to history
    do_run()

    for filename, df_raw in all_dfs.items():
        st.markdown(f'<div class="vvl-card vvl-card-crimson" style="display:flex;align-items:center;gap:12px;margin:1.5rem 0 1rem 0;"><span style="font-size:1.2rem;">📄</span><div><div style="font-weight:700;font-family:\'Cinzel\',serif;">{filename}</div><div style="color:#5A6070;font-size:.82rem;">{df_raw.shape[0]:,} rows · {df_raw.shape[1]} cols</div></div></div>', unsafe_allow_html=True)

        t1,t2,t3,t4,t5 = st.tabs(["🔍 AI Analysis","🧹 Clean Data","📐 DAX Formulas","🔄 Power Query","📊 Dashboard"])

        # ── Analysis ──────────────────────────────────────────────────
        with t1:
            with st.spinner("AI is reading your data..."):
                try:
                    ai_result = ai_analyze(filename, _sum(df_raw))
                    st.session_state[f'ai_{filename}'] = ai_result
                    _add_report_history(email, filename, ai_result.get('domain','General'), df_raw.shape[0])
                except Exception as e:
                    st.error(f"Analysis failed: {e}"); st.stop()

            ai = ai_result
            c1,c2 = st.columns([3,1])
            with c1:
                st.markdown('<span class="vvl-label">What is this data?</span>', unsafe_allow_html=True)
                st.markdown(f'<div class="vvl-card vvl-card-crimson"><div style="font-size:.93rem;line-height:1.7;color:#B0A899;">{ai.get("data_description","")}</div></div>', unsafe_allow_html=True)
            with c2:
                st.markdown('<span class="vvl-label">Domain</span>', unsafe_allow_html=True)
                st.markdown(f'<div class="vvl-card" style="text-align:center;padding:1rem;"><div style="font-size:.68rem;color:#5A6070;text-transform:uppercase;letter-spacing:1px;">Detected</div><div style="font-family:\'Cinzel\',serif;font-size:1.05rem;font-weight:700;color:#F5D76E;margin-top:5px;">{ai.get("domain","General")}</div></div>', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            qa, kb = st.columns(2)
            with qa:
                st.markdown('<span class="vvl-label">Quality Issues</span>', unsafe_allow_html=True)
                iss = ai.get('quality_issues',[])
                if iss:
                    for i in iss: st.markdown(f'<div class="vvl-card vvl-card-warn" style="font-size:.87rem;color:#D4A017;">⚠  {i}</div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div class="vvl-card vvl-card-success" style="font-size:.87rem;">✓  No major quality issues found</div>', unsafe_allow_html=True)
            with kb:
                st.markdown('<span class="vvl-label">Recommended KPIs</span>', unsafe_allow_html=True)
                for kpi in ai.get('kpis',[]): st.markdown(f'<div class="vvl-card vvl-card-gold" style="font-size:.85rem;">🎯  {kpi}</div>', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<span class="vvl-label">Business Insights</span>', unsafe_allow_html=True)
            for ins in ai.get('insights',[]): st.markdown(f'<div class="vvl-card vvl-card-crimson" style="font-size:.87rem;">💡  {ins}</div>', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<span class="vvl-label">Column Summary</span>', unsafe_allow_html=True)
            st.dataframe(pd.DataFrame({'Column':df_raw.columns,'Type':df_raw.dtypes.astype(str).values,'Non-Null':df_raw.count().values,'Missing':df_raw.isnull().sum().values,'Unique':df_raw.nunique().values}), use_container_width=True, hide_index=True)

        # ── Clean ─────────────────────────────────────────────────────
        with t2:
            with st.spinner("Cleaning your data..."):
                df_c, acts = clean(df_raw)
                st.session_state[f'clean_{filename}'] = (df_c, acts)

            m1,m2,m3,m4 = st.columns(4)
            m1.metric("Rows Before",f"{df_raw.shape[0]:,}"); m2.metric("Rows After",f"{df_c.shape[0]:,}")
            m3.metric("Missing Before",f"{df_raw.isnull().sum().sum():,}"); m4.metric("Missing After",f"{df_c.isnull().sum().sum():,}")
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<span class="vvl-label">Actions Taken</span>', unsafe_allow_html=True)
            for a in acts: st.markdown(f'<div class="vvl-card vvl-card-success" style="font-size:.87rem;">✓  {a}</div>', unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            st.dataframe(df_c, use_container_width=True)

        # ── DAX ───────────────────────────────────────────────────────
        with t3:
            ai = st.session_state.get(f'ai_{filename}',{})
            df_c, acts = st.session_state.get(f'clean_{filename}',(df_raw,[]))
            with st.spinner("Generating DAX formulas..."):
                try: dax, tbl = ai_dax(filename, df_c, ai); st.session_state[f'dax_{filename}'] = (dax, tbl)
                except Exception as e: st.error(f"DAX failed: {e}"); dax,tbl={},filename

            st.markdown(f'<div class="vvl-card" style="margin-bottom:1rem;font-size:.88rem;">Power BI table: <code style="color:#F5D76E;">{tbl}</code> &nbsp;·&nbsp; <span style="color:#5A6070;">Modeling → New Measure → Paste any formula</span></div>', unsafe_allow_html=True)
            for name, formula in dax.items():
                with st.expander(f"📐  {name}"):
                    st.markdown(f'<div class="dax-block">{formula}</div>', unsafe_allow_html=True)
                    st.code(formula, language="sql")

        # ── Power Query ───────────────────────────────────────────────
        with t4:
            df_c, acts = st.session_state.get(f'clean_{filename}',(df_raw,[]))
            with st.spinner("Generating Power Query M code..."):
                try: pq = ai_pq(filename, df_raw, df_c, acts); st.session_state[f'pq_{filename}'] = pq
                except Exception as e: pq = f"// Generation failed: {e}"

            st.markdown('<div class="vvl-card vvl-card-gold" style="margin-bottom:1rem;"><div style="color:#D4A017;font-weight:600;margin-bottom:4px;">How to use in Power BI</div><div style="color:#5A6070;font-size:.87rem;line-height:1.8;">Home → Transform Data  →  View → Advanced Editor  →  Delete all → Paste → Update file path → Done</div></div>', unsafe_allow_html=True)
            st.code(pq, language="javascript")

        # ── Charts ────────────────────────────────────────────────────
        with t5:
            ai = st.session_state.get(f'ai_{filename}',{})
            df_c, _ = st.session_state.get(f'clean_{filename}',(df_raw,[]))
            with st.spinner("Building charts..."):
                figs = charts(df_c, ai)
            if figs:
                for fig in figs: st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("Not enough structured data for automatic charts.")

        # ── Downloads ─────────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<span class="vvl-label">Export Your Outputs</span>', unsafe_allow_html=True)
        ai = st.session_state.get(f'ai_{filename}',{})
        df_c, acts = st.session_state.get(f'clean_{filename}',(df_raw,[]))
        dax, tbl = st.session_state.get(f'dax_{filename}',({},filename))
        pq = st.session_state.get(f'pq_{filename}','')
        base = re.sub(r'[^a-zA-Z0-9_]','_',filename.replace('.xlsx','').replace('.csv','')); ts = datetime.now().strftime('%Y%m%d_%H%M')

        d1,d2,d3,d4 = st.columns(4)
        with d1:
            try: st.download_button("📁  Cleaned Excel", data=build_excel(df_c,ai,dax,acts,filename), file_name=f"VVL_{base}_{ts}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            except Exception as e: st.error(f"Excel: {e}")
        with d2:
            try: st.download_button("📄  PDF Report", data=build_pdf(df_c,ai,dax,acts,filename), file_name=f"VVL_Report_{base}_{ts}.pdf", mime="application/pdf", use_container_width=True)
            except Exception as e: st.error(f"PDF: {e}")
        with d3:
            dtxt=f"DAX — {filename}\nTable: '{tbl}'\n\nHOW TO USE: Power BI → Modeling → New Measure → Paste\n\n{'='*55}\n\n"
            for n,f in dax.items(): dtxt+=f"// {n}\n{f}\n\n{'-'*45}\n\n"
            st.download_button("📐  DAX Formulas", data=dtxt, file_name=f"VVL_DAX_{base}_{ts}.txt", mime="text/plain", use_container_width=True)
        with d4:
            st.download_button("🔄  Power Query", data=f"Power Query M — {filename}\n\nHOW TO USE: Power BI → Transform Data → Advanced Editor → Paste\n\n{'='*55}\n\n{pq}", file_name=f"VVL_PQ_{base}_{ts}.txt", mime="text/plain", use_container_width=True)

        st.markdown("<hr>", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# ░░  PAGE: REPORT HISTORY  ░░
# ════════════════════════════════════════════════════════════════

def page_history():
    u = _user(); history = u.get('report_history',[])

    st.markdown('<h1 style="font-family:\'Cinzel\',serif;font-size:1.8rem;font-weight:700;background:linear-gradient(90deg,#F5D76E,#C41E3A);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin:0 0 .4rem 0;">📋 Report History</h1>', unsafe_allow_html=True)
    st.markdown(f'<div style="color:#5A6070;font-size:.9rem;margin-bottom:1.5rem;">All analyses you have run — {len(history)} total</div>', unsafe_allow_html=True)

    if not history:
        st.markdown('<div class="vvl-card" style="text-align:center;padding:3rem;"><div style="font-size:2rem;margin-bottom:.8rem;">📂</div><div style="color:#5A6070;">No reports yet. Head to the BI Analyzer to get started.</div></div>', unsafe_allow_html=True)
        if st.button("Go to BI Analyzer →", key="hist_go"):
            st.session_state['vvl_page'] = 'BI Analyzer'; st.rerun()
        return

    # Summary stats
    domains = {}
    for h in history:
        d = h.get('domain','Other')
        domains[d] = domains.get(d,0)+1

    s1,s2,s3,s4 = st.columns(4)
    s1.metric("Total Reports", len(history))
    s2.metric("Top Domain", max(domains,key=domains.get) if domains else "—")
    s3.metric("Total Rows Processed", f"{sum(h.get('rows',0) for h in history):,}")
    s4.metric("Last Report", history[0].get('date','—') if history else "—")

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<span class="vvl-label">All Reports</span>', unsafe_allow_html=True)

    dom_colors = {"Sales":"#10B981","Finance":"#F5D76E","HR":"#60A5FA","Inventory":"#F59E0B","Marketing":"#A78BFA","Operations":"#F87171"}
    for h in history:
        dc = dom_colors.get(h.get('domain',''),"#C41E3A")
        st.markdown(f"""
        <div class="hist-row">
            <div style="display:flex;align-items:center;gap:12px;">
                <div style="width:10px;height:10px;border-radius:50%;background:{dc};flex-shrink:0;box-shadow:0 0 8px {dc}44;"></div>
                <div>
                    <div style="font-weight:600;font-size:.9rem;color:#E8E2D9;">{h.get('filename','')}</div>
                    <div style="font-size:.75rem;color:#5A6070;">Report ID: {h.get('id','')} · {h.get('rows',0):,} rows</div>
                </div>
            </div>
            <div style="display:flex;align-items:center;gap:12px;">
                <span style="font-size:.75rem;color:{dc};background:rgba(255,255,255,.04);padding:3px 12px;border-radius:99px;border:1px solid rgba(255,255,255,.07);">{h.get('domain','')}</span>
                <span style="font-size:.78rem;color:#5A6070;">{h.get('date','')}</span>
                <span style="font-size:.75rem;color:#10B981;background:rgba(16,185,129,.08);padding:2px 10px;border-radius:99px;">✓ {h.get('status','complete')}</span>
            </div>
        </div>""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# ░░  PAGE: USAGE & BILLING  ░░
# ════════════════════════════════════════════════════════════════

def page_billing():
    u = _user(); plan = u.get('plan','free')
    used = u.get('reports_used',0); lim = u.get('reports_limit',3)
    pct = min(int(used/max(lim,1)*100),100) if lim<999999 else 0
    bar_color = "#EF4444" if pct>80 else "#F59E0B" if pct>55 else "#10B981"

    st.markdown('<h1 style="font-family:\'Cinzel\',serif;font-size:1.8rem;font-weight:700;background:linear-gradient(90deg,#F5D76E,#C41E3A);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin:0 0 .4rem 0;">📊 Usage &amp; Billing</h1>', unsafe_allow_html=True)
    st.markdown('<div style="color:#5A6070;font-size:.9rem;margin-bottom:1.5rem;">Your subscription details and usage statistics</div>', unsafe_allow_html=True)

    left, right = st.columns(2)

    with left:
        st.markdown('<span class="vvl-label">Current Plan</span>', unsafe_allow_html=True)
        p = PLANS.get(plan,{})
        st.markdown(f"""
        <div class="vvl-card vvl-card-crimson" style="padding:1.5rem;">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:1rem;">
                <div style="font-family:'Cinzel',serif;font-size:1.2rem;font-weight:700;color:#E8E2D9;">{p.get('name','')}</div>
                <span class="plan-badge badge-{plan}">{p.get('name','')}</span>
            </div>
            <div style="font-size:2rem;font-weight:700;color:{p.get('color','#E8E2D9')};margin-bottom:.3rem;">
                {"Free" if p.get('price_inr',0)==0 and plan not in ['enterprise'] else ("Custom" if plan=='enterprise' else f"₹{p.get('price_inr',0):,}/mo")}
            </div>
            <div style="color:#5A6070;font-size:.85rem;margin-bottom:1rem;">{p.get('reports',0) if p.get('reports',0)<999999 else "Unlimited"} reports · {p.get('team_seats',1)} team seat{"s" if p.get('team_seats',1)>1 else ""}</div>
            <div style="background:var(--card2);border-radius:8px;padding:.8rem;font-size:.8rem;color:#5A6070;">
                Member since: {u.get('joined','')}
            </div>
        </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<span class="vvl-label">Usage This Cycle</span>', unsafe_allow_html=True)
        st.markdown(f"""
        <div class="vvl-card">
            <div style="display:flex;justify-content:space-between;margin-bottom:10px;">
                <span style="font-size:.85rem;color:#5A6070;">Reports used</span>
                <span style="font-size:.85rem;font-weight:700;color:#E8E2D9;">{used} / {lim if lim<999999 else "∞"}</span>
            </div>
            <div class="usage-bar-bg"><div class="usage-bar-fill" style="background:{bar_color};width:{pct}%;"></div></div>
            <div style="font-size:.75rem;color:#5A6070;margin-top:6px;">{pct}% of your monthly allowance used</div>
        </div>""", unsafe_allow_html=True)

    with right:
        st.markdown('<span class="vvl-label">Upgrade Your Plan</span>', unsafe_allow_html=True)
        for pk in ['starter','pro','enterprise']:
            p2 = PLANS[pk]; is_current = pk == plan
            bc = "vvl-card-crimson" if is_current else ""
            badge = f'<span class="plan-badge badge-{pk}" style="float:right;">{("CURRENT" if is_current else "")}</span>' if is_current else ''
            st.markdown(f"""
            <div class="vvl-card {bc}" style="margin-bottom:.5rem;">
                <div style="display:flex;justify-content:space-between;align-items:center;">
                    <div>
                        <div style="font-family:'Cinzel',serif;font-weight:600;color:#E8E2D9;font-size:.92rem;">{p2['name']}</div>
                        <div style="font-size:.78rem;color:#5A6070;margin-top:2px;">{p2['reports'] if p2['reports']<999999 else "Unlimited"} reports · {p2['team_seats']} seats</div>
                    </div>
                    <div style="text-align:right;">
                        <div style="font-size:1.1rem;font-weight:700;color:{p2['color']}">{f"₹{p2['price_inr']:,}" if p2['price_inr']>0 else "Custom"}<span style="font-size:.72rem;color:#5A6070;">/mo</span></div>
                        {badge}
                    </div>
                </div>
            </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="vvl-card vvl-card-gold" style="padding:1.2rem;">', unsafe_allow_html=True)
        st.markdown('<div style="font-family:\'Cinzel\',serif;font-size:.9rem;color:#D4A017;margin-bottom:.5rem;">Ready to upgrade?</div>', unsafe_allow_html=True)
        st.markdown('<div style="font-size:.82rem;color:#5A6070;margin-bottom:.8rem;">Contact us to upgrade your plan. We accept UPI and bank transfer.</div>', unsafe_allow_html=True)
        st.markdown('**📧 admin@vampirevenomlabs.com**', unsafe_allow_html=False)
        st.markdown('</div>', unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# ░░  PAGE: TEAM ACCESS  ░░
# ════════════════════════════════════════════════════════════════

def page_team():
    u = _user(); plan = u.get('plan','free'); email = st.session_state.get('vvl_email','')
    max_seats = PLANS.get(plan,{}).get('team_seats',1)
    db = _db()

    st.markdown('<h1 style="font-family:\'Cinzel\',serif;font-size:1.8rem;font-weight:700;background:linear-gradient(90deg,#F5D76E,#C41E3A);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin:0 0 .4rem 0;">👥 Team Access</h1>', unsafe_allow_html=True)
    st.markdown('<div style="color:#5A6070;font-size:.9rem;margin-bottom:1.5rem;">Invite colleagues to share your subscription</div>', unsafe_allow_html=True)

    if max_seats <= 1 and plan not in ['admin']:
        st.markdown(f"""<div class="vvl-card" style="text-align:center;padding:2rem;border-color:rgba(196,30,58,.3);">
        <div style="font-size:1.5rem;margin-bottom:.6rem;">👥</div>
        <div style="font-family:'Cinzel',serif;color:#F5D76E;margin-bottom:.4rem;">Team Access — Pro Feature</div>
        <div style="color:#5A6070;font-size:.88rem;">Your current plan ({PLANS.get(plan,{}).get('name','')}) supports 1 seat.<br>
        Upgrade to Starter (2 seats) or Pro (5 seats) to add team members.</div>
        </div>""", unsafe_allow_html=True)
        if st.button("💎 Upgrade to Add Team Members", key="team_upgrade"):
            st.session_state['vvl_page'] = 'Upgrade Plan'; st.rerun()
        return

    # Get or create team
    team_id = u.get('team_id')
    if not team_id:
        team_id = f"team-{email[:8]}-{str(uuid.uuid4())[:4]}"
        db['users'][email]['team_id'] = team_id
        db['teams'][team_id] = {"owner": email, "members": [email], "created": datetime.now().strftime('%Y-%m-%d')}
        u = _user()

    team = db.get('teams',{}).get(team_id, {"owner": email, "members": [email]})
    members = team.get('members', [email])
    is_owner = team.get('owner','') == email

    m1,m2,m3 = st.columns(3)
    m1.metric("Team Seats", max_seats); m2.metric("Used Seats", len(members)); m3.metric("Available", max(max_seats-len(members),0))

    st.markdown("<br>", unsafe_allow_html=True)
    left, right = st.columns([1.5,1])

    with left:
        st.markdown('<span class="vvl-label">Team Members</span>', unsafe_allow_html=True)
        for m in members:
            mu = db['users'].get(m,{})
            is_you = m == email
            st.markdown(f"""
            <div class="team-row">
                <div style="display:flex;align-items:center;gap:10px;">
                    <div style="width:32px;height:32px;border-radius:50%;background:linear-gradient(135deg,#C41E3A,#8B0000);display:flex;align-items:center;justify-content:center;font-size:.85rem;font-weight:700;color:#F5D76E;flex-shrink:0;">{mu.get('name','?')[0].upper()}</div>
                    <div>
                        <div style="font-weight:600;font-size:.88rem;color:#E8E2D9;">{mu.get('name','Unknown')}{' (you)' if is_you else ''}</div>
                        <div style="font-size:.75rem;color:#5A6070;">{m}</div>
                    </div>
                </div>
                <div>
                    <span style="font-size:.72rem;color:{'#D4A017' if m==team.get('owner') else '#5A6070'};background:rgba(255,255,255,.04);padding:2px 10px;border-radius:99px;">
                        {'Owner' if m==team.get('owner') else 'Member'}
                    </span>
                </div>
            </div>""", unsafe_allow_html=True)

    with right:
        if is_owner and len(members) < max_seats:
            st.markdown('<span class="vvl-label">Invite Member</span>', unsafe_allow_html=True)
            inv_email = st.text_input("Colleague's email", placeholder="colleague@company.com", key="inv_email")
            if st.button("Send Invite →", key="inv_btn"):
                inv_email = inv_email.strip().lower()
                if inv_email in db['users']:
                    if inv_email not in members:
                        members.append(inv_email)
                        db['teams'][team_id]['members'] = members
                        db['users'][inv_email]['team_id'] = team_id
                        st.success(f"✓ {inv_email} added to your team!")
                        st.rerun()
                    else:
                        st.warning("Already a team member")
                else:
                    st.error("That email has no account. Ask them to register first.")
        elif len(members) >= max_seats:
            st.markdown(f'<div class="vvl-card vvl-card-warn" style="font-size:.85rem;color:#D4A017;padding:1rem;">All {max_seats} seats filled. Upgrade your plan to add more members.</div>', unsafe_allow_html=True)
            if st.button("💎 Upgrade Seats", key="seats_upgrade"):
                st.session_state['vvl_page'] = 'Upgrade Plan'; st.rerun()

        if is_owner and len(members) > 1:
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<span class="vvl-label">Remove Member</span>', unsafe_allow_html=True)
            non_owners = [m for m in members if m != email]
            rm = st.selectbox("Select member to remove", non_owners, key="rm_sel")
            if st.button("Remove", key="rm_btn"):
                members.remove(rm); db['teams'][team_id]['members'] = members
                db['users'][rm]['team_id'] = None
                st.success(f"✓ Removed {rm}"); st.rerun()


# ════════════════════════════════════════════════════════════════
# ░░  PAGE: UPGRADE  ░░
# ════════════════════════════════════════════════════════════════

def page_upgrade():
    u = _user(); current = u.get('plan','free')
    st.markdown('<h1 style="font-family:\'Cinzel\',serif;font-size:1.8rem;font-weight:700;background:linear-gradient(90deg,#F5D76E,#C41E3A);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin:0 0 .4rem 0;">💎 Upgrade Your Plan</h1>', unsafe_allow_html=True)
    st.markdown('<div style="color:#5A6070;font-size:.9rem;margin-bottom:2rem;">Choose the plan that fits your needs</div>', unsafe_allow_html=True)

    cols = st.columns(4)
    for col, pk in zip(cols, ['free','starter','pro','enterprise']):
        p = PLANS[pk]; is_cur = pk==current; is_pop = pk=='pro'
        bdr = "border:2px solid rgba(212,160,23,.5);box-shadow:0 0 30px rgba(212,160,23,.12);" if is_pop else ("border:2px solid rgba(196,30,58,.4);" if is_cur else "")
        with col:
            st.markdown(f"""
            <div class="vvl-card" style="{bdr}text-align:center;padding:1.8rem 1rem;min-height:280px;">
                {'<div style="font-size:.68rem;color:#D4A017;font-weight:700;letter-spacing:1px;margin-bottom:.4rem;">⭐ MOST POPULAR</div>' if is_pop else ('<div style="font-size:.68rem;color:#C41E3A;font-weight:700;letter-spacing:1px;margin-bottom:.4rem;">✓ CURRENT PLAN</div>' if is_cur else '<div style="height:1.1rem;"></div>')}
                <div style="font-family:'Cinzel',serif;font-size:.95rem;font-weight:700;color:#E8E2D9;margin-bottom:.4rem;">{p['name']}</div>
                <div style="font-size:1.9rem;font-weight:700;color:{p['color']};margin:.6rem 0;">
                    {"Free" if p['price_inr']==0 and pk!='enterprise' else ("Custom" if pk=='enterprise' else f"₹{p['price_inr']:,}")}
                    <span style="font-size:.72rem;color:#5A6070;">{"" if pk in ["free","enterprise"] else "/mo"}</span>
                </div>
                <div style="font-size:.82rem;color:#5A6070;margin-bottom:.4rem;">{p['reports'] if p['reports']<999999 else "Unlimited"} reports/month</div>
                <div style="font-size:.78rem;color:#3D4455;">{p['team_seats']} team seat{"s" if p['team_seats']>1 else ""}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("""
    <div class="vvl-card vvl-card-gold" style="max-width:600px;margin:0 auto;text-align:center;padding:2rem;">
        <div style="font-family:'Cinzel',serif;font-size:1.05rem;color:#D4A017;margin-bottom:.6rem;">How to Upgrade</div>
        <div style="font-size:.88rem;color:#5A6070;line-height:1.8;margin-bottom:1rem;">
            1. Choose your plan above<br>
            2. Pay via UPI / Bank Transfer / Razorpay<br>
            3. Email your payment receipt<br>
            4. We activate your plan within 2 hours
        </div>
        <div style="font-size:.95rem;color:#E8E2D9;font-weight:600;">📧 admin@vampirevenomlabs.com</div>
        <div style="font-size:.82rem;color:#5A6070;margin-top:.4rem;">UPI: admin@vvl (example)</div>
    </div>""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# ░░  PAGE: ADMIN PANEL  ░░
# ════════════════════════════════════════════════════════════════

def page_admin():
    if not is_admin():
        st.error("Access denied."); return

    db = _db()
    st.markdown('<h1 style="font-family:\'Cinzel\',serif;font-size:1.8rem;font-weight:700;background:linear-gradient(90deg,#F87171,#C41E3A);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin:0 0 .4rem 0;">⚙️ Admin Panel</h1>', unsafe_allow_html=True)
    st.markdown('<div style="color:#5A6070;font-size:.9rem;margin-bottom:1.5rem;">Platform management — all users, API config, stats</div>', unsafe_allow_html=True)

    t_api, t_users, t_stats = st.tabs(["🔑 API Configuration","👤 User Management","📈 Platform Stats"])

    # ── API Config ────────────────────────────────────────────────
    with t_api:
        env_key = os.environ.get("CLAUDE_API_KEY","")
        try:
            sec_key = st.secrets.get("CLAUDE_API_KEY","")
        except Exception:
            sec_key = ""
        override = db.get('admin_api_override','')
        active_key = override if (override and len(override)>20) else (sec_key or env_key)

        st.markdown('<span class="vvl-label">Current API Key Status</span>', unsafe_allow_html=True)
        if active_key:
            src = "Admin Override" if override and len(override)>20 else ("Streamlit Secrets" if sec_key else "Environment Variable")
            st.markdown(f'<div class="vvl-card vvl-card-success"><span style="color:#10B981;">✓ API key configured</span> &nbsp;·&nbsp; Source: <b>{src}</b> &nbsp;·&nbsp; Ends in: <code style="color:#F5D76E;">...{active_key[-6:]}</code></div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="vvl-card" style="border-left:3px solid #EF4444;color:#EF4444;">✗ No API key found. Set one below or in Streamlit secrets.</div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<span class="vvl-label">Admin Override Key</span>', unsafe_allow_html=True)
        st.markdown('<div style="font-size:.82rem;color:#5A6070;margin-bottom:.6rem;">Paste a key here to override the environment variable. Leave blank to use env/secrets key. This key is NEVER shown to users.</div>', unsafe_allow_html=True)
        new_override = st.text_input("Override API Key", value=override, type="password", placeholder="sk-ant-api03-...", key="admin_api_input")
        c1,c2,_ = st.columns([1,1,2])
        with c1:
            if st.button("💾 Save Override", key="save_api"):
                db['admin_api_override'] = new_override.strip()
                st.success("✓ Saved! Override active." if new_override.strip() else "✓ Cleared — using env/secrets key.")
        with c2:
            if st.button("🗑 Clear Override", key="clear_api"):
                db['admin_api_override'] = ""
                st.success("✓ Cleared — now using env/secrets key.")

        st.markdown("<br>", unsafe_allow_html=True)
        with st.expander("📖 How to set API key in Streamlit Cloud (recommended)"):
            st.markdown("""
            1. Go to **Streamlit Cloud** → Your App → **⋮ Menu** → **Settings**
            2. Click **Secrets** tab
            3. Add exactly:
            ```toml
            CLAUDE_API_KEY = "sk-ant-your-key-here"
            ```
            4. Click **Save** — app restarts, key is active
            5. Users **never** see this key — it only runs on the server
            """)

    # ── User Management ───────────────────────────────────────────
    with t_users:
        users = db.get('users',{})
        st.markdown('<span class="vvl-label">All Users</span>', unsafe_allow_html=True)

        user_rows = []
        for em, u in users.items():
            user_rows.append({"Email":em,"Name":u.get('name',''),"Plan":u.get('plan',''),"Used":u.get('reports_used',0),"Limit":u.get('reports_limit',0),"Joined":u.get('joined',''),"Active":u.get('active',True)})
        st.dataframe(pd.DataFrame(user_rows), use_container_width=True, hide_index=True)

        st.markdown("<br>", unsafe_allow_html=True)
        with st.expander("➕ Add / Update User"):
            c1,c2 = st.columns(2)
            with c1:
                ae = st.text_input("Email", key="adm_em")
                an = st.text_input("Full Name", key="adm_nm")
                ap = st.text_input("Password (blank = keep existing)", type="password", key="adm_pw")
            with c2:
                apl = st.selectbox("Plan", list(PLANS.keys()), key="adm_pl")
                alm = st.number_input("Report Limit", 0, 999999, value=PLANS[st.session_state.get('adm_pl','free')]['reports'], key="adm_lm")
                aac = st.checkbox("Account Active", value=True, key="adm_ac")
            if st.button("💾 Save User", key="adm_save"):
                if ae.strip():
                    ek = ae.strip().lower()
                    if ek in users:
                        users[ek]['plan']=apl; users[ek]['reports_limit']=alm; users[ek]['active']=aac
                        if an: users[ek]['name']=an
                        if ap: users[ek]['password_hash']=_hash(ap)
                        st.success(f"✓ Updated {ek}")
                    else:
                        if an and ap:
                            users[ek]={"uid":str(uuid.uuid4())[:8],"name":an,"password_hash":_hash(ap),"plan":apl,"reports_used":0,"reports_limit":alm,"joined":datetime.now().strftime('%Y-%m-%d'),"active":aac,"team_id":None,"report_history":[]}
                            st.success(f"✓ Created account for {ek}")
                        else:
                            st.error("For new users: provide name + password")
                    st.rerun()

    # ── Stats ─────────────────────────────────────────────────────
    with t_stats:
        users = db.get('users',{})
        s1,s2,s3,s4 = st.columns(4)
        s1.metric("Total Users", len(users))
        s2.metric("Reports Run", sum(u.get('reports_used',0) for u in users.values()))
        s3.metric("Active Accounts", sum(1 for u in users.values() if u.get('active',True)))
        s4.metric("Paying Users", sum(1 for u in users.values() if u.get('plan') not in ['free','admin']))

        st.markdown("<br>", unsafe_allow_html=True)
        plan_counts = {}
        for u in users.values():
            p = u.get('plan','free'); plan_counts[p] = plan_counts.get(p,0)+1
        plan_df = pd.DataFrame({"Plan": list(plan_counts.keys()), "Count": list(plan_counts.values())})
        st.markdown('<span class="vvl-label">Users by Plan</span>', unsafe_allow_html=True)
        st.dataframe(plan_df, use_container_width=True, hide_index=True)


# ════════════════════════════════════════════════════════════════
# ░░  MAIN ROUTER  ░░
# ════════════════════════════════════════════════════════════════

def main():
    # Initialise state
    if 'vvl_logged_in' not in st.session_state:
        st.session_state['vvl_logged_in'] = False
    if 'vvl_page' not in st.session_state:
        st.session_state['vvl_page'] = 'Dashboard'
    _ = _db()   # ensure DB is initialised

    # Not logged in → show auth page
    if not st.session_state['vvl_logged_in']:
        page_auth()
        return

    # Logged in → sidebar + routed page
    render_sidebar()

    page = st.session_state.get('vvl_page','Dashboard')
    page_map = {
        'Dashboard':     page_dashboard,
        'BI Analyzer':   page_analyzer,
        'Report History': page_history,
        'Usage & Billing': page_billing,
        'Team Access':   page_team,
        'Upgrade Plan':  page_upgrade,
        'Admin Panel':   page_admin,
        'My Reports':    page_history,   # alias
    }
    page_map.get(page, page_dashboard)()

main()
